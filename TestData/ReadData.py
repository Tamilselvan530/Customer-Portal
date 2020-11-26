import openpyxl


workbook=openpyxl.load_workbook("C:/Users/SS372669/PycharmProjects/Customer_Portal/TestData/Data.xlsx")
sheet=workbook["Sheet1"]


def fetch_number_of_rows(Sheetname):
    sh = workbook[Sheetname]
    return sh.max_row


def fetch_cell_data(Sheetname, row, column):
    sh = workbook[Sheetname]
    cell = sh.cell(int(row), int(column))
    print(cell.value)
    return cell.value


#print(fetch_number_of_rows("Sheet1"))
print(fetch_cell_data("Sheet1", 2, 2))
